import csv
import os
import sqlite3

from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook

DB_NAME = "attendance.db"
if os.environ.get("RESET_DB_ON_START") == "1":
    if os.path.exists(DB_NAME):
        os.remove(DB_NAME)

CSV_FILE = "students.csv"

EXCEL_ROOM_FILES = {
    "Neural": "Neural.xlsx",
    "Qubit": "Qubit.xlsx",
    "Quantum Core": "QuantumCore.xlsx",
    "Intelligence": "Intelligence.xlsx",
}

VALID_ROOMS = tuple(EXCEL_ROOM_FILES.keys())

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")


def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def ensure_column_exists(conn, table_name: str, column_name: str, column_def: str):
    cols = conn.execute(f"PRAGMA table_info({table_name});").fetchall()
    existing = {c["name"] for c in cols}
    if column_name not in existing:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_def};")


def init_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("PRAGMA foreign_keys = ON;")

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS teams (
            id INTEGER PRIMARY KEY,
            team_name TEXT NOT NULL,
            room TEXT NOT NULL
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            team_id INTEGER NOT NULL,
            checked_in INTEGER NOT NULL DEFAULT 0,
            checkin_time TEXT,
            university TEXT,
            FOREIGN KEY (team_id) REFERENCES teams (id)
        );
        """
    )

    ensure_column_exists(conn, "students", "checked_in", "checked_in INTEGER NOT NULL DEFAULT 0")
    ensure_column_exists(conn, "students", "checkin_time", "checkin_time TEXT")
    ensure_column_exists(conn, "students", "university", "university TEXT")

    cur.execute("SELECT COUNT(*) AS c FROM students;")
    count = cur.fetchone()["c"]

    if count == 0:
        loaded = try_load_from_excels(conn)
        if not loaded:
            if os.path.exists(CSV_FILE):
                load_from_csv(conn, CSV_FILE)


    conn.commit()
    conn.close()


def clean_cell(x):
    if x is None:
        return ""
    return str(x).strip()


def insert_team_if_needed(conn, team_id: int, team_name: str, room: str):
    conn.execute(
        """
        INSERT OR IGNORE INTO teams (id, team_name, room)
        VALUES (?, ?, ?)
        """,
        (team_id, team_name, room),
    )


def insert_student_if_needed(conn, student_name: str, team_id: int, university: str):
    conn.execute(
        """
        INSERT INTO students (name, team_id, university)
        SELECT ?, ?, ?
        WHERE NOT EXISTS (
            SELECT 1 FROM students WHERE name = ? AND team_id = ?
        )
        """,
        (student_name, team_id, university or None, student_name, team_id),
    )


def try_load_from_excels(conn) -> bool:
    any_file = False

    for room, path in EXCEL_ROOM_FILES.items():
        if not os.path.exists(path):
            continue

        any_file = True

        wb = load_workbook(path, data_only=True)

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 1:
                    continue

                team_raw = row[0]
                if team_raw is None:
                    continue

                try:
                    team_id = int(str(team_raw).strip())
                except Exception:
                    continue

                name1 = clean_cell(row[1]) if len(row) > 1 else ""
                uni1 = clean_cell(row[2]) if len(row) > 2 else ""
                name2 = clean_cell(row[3]) if len(row) > 3 else ""
                uni2 = clean_cell(row[4]) if len(row) > 4 else ""

                team_name = clean_cell(row[5]) if len(row) > 5 else ""
                if not team_name:
                    team_name = f"فريق {team_id}"

                insert_team_if_needed(conn, team_id, team_name, room)

                if name1:
                    insert_student_if_needed(conn, name1, team_id, uni1)
                if name2:
                    insert_student_if_needed(conn, name2, team_id, uni2)

        conn.commit()




def load_from_csv(conn, csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            team_id = int(row["team_id"])
            team_name = (row.get("team_name") or "").strip() or f"فريق {team_id}"
            room = (row.get("room") or "").strip() or "Neural"
            student_name = (row.get("student_name") or "").strip()
            university = (row.get("university") or "").strip()

            if room not in VALID_ROOMS:
                room = "Neural"

            insert_team_if_needed(conn, team_id, team_name, room)
            if student_name:
                insert_student_if_needed(conn, student_name, team_id, university)

    conn.commit()



init_db()


@app.route("/")
def index():
    return redirect(url_for("checkin"))


@app.route("/checkin", methods=["GET", "POST"])
def checkin():
    conn = get_db_connection()

    if request.method == "POST":
        student_id = request.form.get("student_id")
        team_id = request.form.get("team_id")
        action = (request.form.get("action") or "checkin").strip().lower()

        if team_id and str(team_id).isdigit():
            team_id_int = int(team_id)

            if action == "checkout":
                cur = conn.execute(
                    """
                    UPDATE students
                    SET checked_in = 0,
                        checkin_time = NULL
                    WHERE team_id = ?
                      AND checked_in = 1
                    """,
                    (team_id_int,),
                )
                conn.commit()

            else:
                cur = conn.execute(
                    """
                    UPDATE students
                    SET checked_in = 1,
                        checkin_time = datetime('now', 'localtime')
                    WHERE team_id = ?
                      AND checked_in = 0
                    """,
                    (team_id_int,),
                )
                conn.commit()


        elif student_id and str(student_id).isdigit():
            student_id_int = int(student_id)

            if action == "checkout":
                cur = conn.execute(
                    """
                    UPDATE students
                    SET checked_in = 0,
                        checkin_time = NULL
                    WHERE id = ?
                      AND checked_in = 1
                    """,
                    (student_id_int,),
                )
                conn.commit()

            else:
                cur = conn.execute(
                    """
                    UPDATE students
                    SET checked_in = 1,
                        checkin_time = datetime('now', 'localtime')
                    WHERE id = ?
                      AND checked_in = 0
                    """,
                    (student_id_int,),
                )
                conn.commit()
        conn.close()
        return redirect(url_for("checkin"))

    q = request.args.get("q", "").strip()
    room = request.args.get("room", "").strip()

    sql = """
        SELECT
            students.id,
            students.name,
            students.university,
            students.checked_in,
            students.checkin_time,
            teams.id AS team_id,
            teams.team_name,
            teams.room
        FROM students
        JOIN teams ON students.team_id = teams.id
        WHERE 1 = 1
    """
    params = []

    if q:
        sql += """
            AND (
                students.name LIKE ?
                OR students.university LIKE ?
                OR teams.team_name LIKE ?
                OR CAST(teams.id AS TEXT) LIKE ?
            )
        """
        like = f"%{q}%"
        params += [like, like, like, like]

    if room in VALID_ROOMS:
        sql += " AND teams.room = ?"
        params.append(room)

    sql += " ORDER BY teams.room, teams.id, students.id"

    students = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template("checkin.html", students=students, q=q, room=room, rooms=VALID_ROOMS)


@app.route("/students")
def students_list():
    conn = get_db_connection()

    status = request.args.get("status", "all")
    room = request.args.get("room", "").strip()

    sql = """
        SELECT
            students.id,
            students.name,
            students.university,
            students.checked_in,
            students.checkin_time,
            teams.id AS team_id,
            teams.team_name,
            teams.room
        FROM students
        JOIN teams ON students.team_id = teams.id
        WHERE 1 = 1
    """
    params = []

    if status == "present":
        sql += " AND students.checked_in = 1"
    elif status == "absent":
        sql += " AND students.checked_in = 0"

    if room in VALID_ROOMS:
        sql += " AND teams.room = ?"
        params.append(room)

    sql += " ORDER BY teams.room, teams.id, students.id"

    students = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template("students_list.html", students=students, status=status, room=room, rooms=VALID_ROOMS)


@app.route("/stats")
def stats():
    conn = get_db_connection()

    total_students = conn.execute("SELECT COUNT(*) AS c FROM students").fetchone()["c"]
    total_checked_in = conn.execute("SELECT COUNT(*) AS c FROM students WHERE checked_in = 1").fetchone()["c"]

    per_room = conn.execute(
        """
        SELECT
            teams.room,
            COUNT(students.id) AS total_students,
            SUM(CASE WHEN students.checked_in = 1 THEN 1 ELSE 0 END) AS present_students
        FROM students
        JOIN teams ON students.team_id = teams.id
        GROUP BY teams.room
        ORDER BY teams.room
        """
    ).fetchall()

    conn.close()

    return render_template(
        "stats.html",
        total_students=total_students,
        total_checked_in=total_checked_in,
        per_room=per_room,
        rooms=VALID_ROOMS,
    )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--reset-db", action="store_true")
    args = parser.parse_args()

    if args.reset_db and os.path.exists(DB_NAME):
        os.remove(DB_NAME)
        init_db()

    app.run(host="0.0.0.0", port=5000, debug=not args.reset_db, use_reloader=not args.reset_db)

    app.run(host="0.0.0.0", port=5000, debug=not args.reset_db, use_reloader=not args.reset_db)

